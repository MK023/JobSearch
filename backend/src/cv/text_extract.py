"""Extract plain text from CV files (PDF, DOCX, DOC, TXT, XLSX, XLS)."""

import subprocess
import tempfile
from io import BytesIO
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader

ALLOWED_EXTENSIONS = {".txt", ".pdf", ".doc", ".docx", ".xlsx", ".xls"}
MAX_FILE_BYTES = 10 * 1024 * 1024  # 10 MB


def extract_text(file_bytes: bytes, filename: str) -> str:
    """Extract plain text from an uploaded file. Raises ValueError on failure."""
    ext = Path(filename).suffix.lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(f"Formato non supportato: {ext}. Usa PDF, DOCX, DOC, TXT o XLSX.")
    if len(file_bytes) > MAX_FILE_BYTES:
        raise ValueError("File troppo grande (max 10 MB).")
    if len(file_bytes) == 0:
        raise ValueError("File vuoto.")

    extractors = {
        ".pdf": _extract_pdf,
        ".docx": _extract_docx,
        ".doc": _extract_doc,
        ".txt": _extract_txt,
        ".xlsx": _extract_xlsx,
        ".xls": _extract_xlsx,
    }
    text = extractors[ext](file_bytes)
    text = text.strip()
    if len(text) < 20:
        raise ValueError("Impossibile estrarre testo sufficiente dal file.")
    return text


def _extract_pdf(data: bytes) -> str:
    reader = PdfReader(BytesIO(data))
    pages = [page.extract_text() or "" for page in reader.pages]
    return "\n".join(pages)


def _extract_docx(data: bytes) -> str:
    doc = Document(BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs)


def _extract_doc(data: bytes) -> str:
    with tempfile.NamedTemporaryFile(suffix=".doc", delete=True) as tmp:
        tmp.write(data)
        tmp.flush()
        result = subprocess.run(  # noqa: S603
            ["antiword", tmp.name],  # noqa: S607
            capture_output=True,
            text=True,
            timeout=10,
        )
        if result.returncode != 0:
            raise ValueError("Impossibile leggere il file DOC. Prova a convertirlo in DOCX.")
        return result.stdout


def _extract_xlsx(data: bytes) -> str:
    """Extract text from all cells across all sheets."""
    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(" | ".join(cells))
    wb.close()
    return "\n".join(lines)


def _extract_txt(data: bytes) -> str:
    for encoding in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            return data.decode(encoding)
        except ValueError:
            # UnicodeDecodeError is a subclass of ValueError — catching
            # ValueError alone covers both malformed-UTF8 and other codec issues.
            continue
    raise ValueError("Encoding del file non riconosciuto. Salva il file come UTF-8.")
